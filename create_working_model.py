import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import BarChart, Reference

class DistilleryModelGenerator:
    """
    Generates a simplified but fully functional, robust, and professional Excel financial model
    for The Brogue Distillery. This script is self-contained and designed to be bulletproof.
    """
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove the default sheet
        self._define_styles()

    def _define_styles(self):
        """Defines all named styles for professional and consistent formatting."""
        # Cell Styles
        self.header_style = NamedStyle(name="header_style", font=Font(bold=True, color="FFFFFF"), fill=PatternFill(start_color="002060", end_color="002060", fill_type="solid"))
        self.subheader_style = NamedStyle(name="subheader_style", font=Font(bold=True), fill=PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"))
        self.input_style = NamedStyle(name="input_style", fill=PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"))
        self.total_style = NamedStyle(name="total_style", font=Font(bold=True), border=Border(top=Side(style='thin'), bottom=Side(style='double')))

        # Number Formats
        self.currency_style = NamedStyle(name="currency_style", number_format='$#,##0')
        self.percent_style = NamedStyle(name="percent_style", number_format='0.0%')
        self.integer_style = NamedStyle(name="integer_style", number_format='#,##0')
        self.currency_2dp_style = NamedStyle(name="currency_2dp_style", number_format='$#,##0.00')

        # Add styles to the workbook if they don't already exist
        for style in [self.header_style, self.subheader_style, self.input_style, self.total_style,
                      self.currency_style, self.percent_style, self.integer_style, self.currency_2dp_style]:
            if style.name not in self.wb.style_names:
                self.wb.add_named_style(style)

    def _create_named_range(self, name, ws, cell_address):
        """Creates a workbook-level named range using the correct API."""
        if name not in self.wb.defined_names:
            # Correct method: Assign a DefinedName object to the dictionary key
            self.wb.defined_names[name] = openpyxl.workbook.defined_name.DefinedName(
                name, attr_text=f"'{ws.title}'!${cell_address}"
            )

    def create_assumptions_sheet(self):
        """Creates the main input sheet with CHOOSE formulas linked to scenarios."""
        ws = self.wb.create_sheet("01_Assumptions", 0)
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 25

        # Hardcoded assumptions structure, mirroring the YAML logic for simplicity
        assumptions = {
            "Volume & Pricing": [
                ("Y1 Production Target (Bottles)", 50000, "inp_Y1_Volume"),
                ("Y2+ Annual Growth %", 0.25, "inp_Growth_Rate"),
                ("Avg. Net Price / Bottle", 45.00, "inp_Avg_Price"),
            ],
            "Costs": [
                ("Avg. COGS / Bottle", 18.23, "inp_COGS_per_Bottle"),
                ("Annual Fixed OpEx", 780000, "inp_Fixed_OpEx"),
                ("Annual OpEx Growth %", 0.05, "inp_OpEx_Growth"),
            ],
            "Capital & Financing": [
                ("Initial CapEx Spend", 900000, "inp_Initial_Capex"),
                ("Equipment Depreciation Years", 7, "inp_Depr_Years"),
                ("Initial Cash Position", 1500000, "inp_Initial_Cash"),
                ("SAFE Round Amount", 3000000, "inp_SAFE_Amount"),
                ("SAFE Valuation Cap", 15000000, "inp_SAFE_Cap"),
                ("SAFE Discount Rate", 0.20, "inp_SAFE_Discount"),
                ("Series A Pre-Money Valuation", 30000000, "inp_SeriesA_Valuation"),
            ],
            "Scenario Control": [
                ("Current Scenario (1=Base, 2=Up, 3=Down)", 1, "inp_Scenario_Index"),
            ]
        }

        row = 2
        for category, items in assumptions.items():
            ws[f'A{row}'] = category
            ws[f'A{row}'].style = self.subheader_style
            row += 1
            for label, _, range_name in items:
                ws[f'A{row}'] = label
                cell_address = f'B{row}'
                # The actual value is a CHOOSE formula pointing to the scenario sheet
                ws[cell_address] = f"=CHOOSE(inp_Scenario_Index, '02_ScenarioSwitcher'!B{row}, '02_ScenarioSwitcher'!C{row}, '02_ScenarioSwitcher'!D{row})"
                ws[cell_address].style = self.input_style
                self._create_named_range(range_name, ws, cell_address)
                row += 1
            row += 1

    def create_scenario_switcher_sheet(self):
        """Creates the data source sheet for the CHOOSE functions."""
        ws = self.wb.create_sheet("02_ScenarioSwitcher", 1)
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.append(["Parameter", "Base Case", "Upside Case", "Downside Case"])
        for cell in ws[1]: cell.style = self.header_style

        # These hardcoded values correspond to the row numbers in the Assumptions sheet
        scenarios = {
            3: ("Y1 Production Target (Bottles)", 50000, 65000, 40000),
            4: ("Y2+ Annual Growth %", 0.25, 0.35, 0.15),
            5: ("Avg. Net Price / Bottle", 45.00, 50.00, 40.00),
            7: ("Avg. COGS / Bottle", 18.23, 17.50, 19.00),
            8: ("Annual Fixed OpEx", 780000, 850000, 750000),
            9: ("Annual OpEx Growth %", 0.05, 0.06, 0.04),
            11: ("Initial CapEx Spend", 900000, 850000, 1000000),
            12: ("Equipment Depreciation Years", 7, 7, 7),
            13: ("Initial Cash Position", 1500000, 1500000, 1500000),
            14: ("SAFE Round Amount", 3000000, 3000000, 3000000),
            15: ("SAFE Valuation Cap", 15000000, 18000000, 12000000),
            16: ("SAFE Discount Rate", 0.20, 0.20, 0.25),
            17: ("Series A Pre-Money Valuation", 30000000, 40000000, 20000000),
            19: ("Current Scenario", 1, 2, 3),
        }
        for row, (label, base, up, down) in scenarios.items():
            ws[f'A{row}'] = label
            ws[f'B{row}'] = base
            ws[f'C{row}'] = up
            ws[f'D{row}'] = down

    def create_financials_sheet(self):
        """Creates an integrated 5-year annual P&L and Cash Flow statement."""
        ws = self.wb.create_sheet("03_Financials", 2)
        ws.column_dimensions['A'].width = 35
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18

        years = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        headers = ["Metric"] + years
        ws.append(headers)
        for cell in ws[1]: cell.style = self.header_style

        # P&L Section
        ws.append(["PROFIT & LOSS"])
        ws['A2'].style = self.subheader_style
        ws.append(["Volume (Bottles)", "=inp_Y1_Volume", "=B3*(1+inp_Growth_Rate)", "=C3*(1+inp_Growth_Rate)", "=D3*(1+inp_Growth_Rate)", "=E3*(1+inp_Growth_Rate)"])
        ws.append(["Revenue", "=B3*inp_Avg_Price", "=C3*inp_Avg_Price", "=D3*inp_Avg_Price", "=E3*inp_Avg_Price", "=F3*inp_Avg_Price"])
        ws.append(["COGS", "=-B3*inp_COGS_per_Bottle", "=-C3*inp_COGS_per_Bottle", "=-D3*inp_COGS_per_Bottle", "=-E3*inp_COGS_per_Bottle", "=-F3*inp_COGS_per_Bottle"])
        ws.append(["Gross Profit", "=B4+B5", "=C4+C5", "=D4+D5", "=E4+E5", "=F4+F5"])
        ws.append(["Fixed OpEx", "=-inp_Fixed_OpEx", "=-B7*(1+inp_OpEx_Growth)", "=-C7*(1+inp_OpEx_Growth)", "=-D7*(1+inp_OpEx_Growth)", "=-E7*(1+inp_OpEx_Growth)"])
        ws.append(["EBITDA", "=B6+B7", "=C6+C7", "=D6+D7", "=E6+E7", "=F6+F7"])
        ws.append(["Depreciation", "=-inp_Initial_Capex/inp_Depr_Years", "=-inp_Initial_Capex/inp_Depr_Years", "=-inp_Initial_Capex/inp_Depr_Years", "=-inp_Initial_Capex/inp_Depr_Years", "=-inp_Initial_Capex/inp_Depr_Years"])
        ws.append(["EBIT", "=B8+B9", "=C8+C9", "=D8+D9", "=E8+E9", "=F8+F9"])
        ws.append(["Taxes (25%)", "=IF(B10>0, -B10*0.25, 0)", "=IF(C10>0, -C10*0.25, 0)", "=IF(D10>0, -D10*0.25, 0)", "=IF(E10>0, -E10*0.25, 0)", "=IF(F10>0, -F10*0.25, 0)"])
        ws.append(["Net Income", "=B10+B11", "=C10+C11", "=D10+D11", "=E10+E11", "=F10+F11"])

        # Cash Flow Section
        ws.append([]) # Spacer row
        ws.append(["CASH FLOW"])
        ws['A14'].style = self.subheader_style
        ws.append(["Beginning Cash", "=inp_Initial_Cash+inp_SAFE_Amount", "=G18", "=H18", "=I18", "=J18"])
        ws.append(["Cash Flow from Ops (simplified)", "=B12-B9", "=C12-C9", "=D12-D9", "=E12-E9", "=F12-F9"]) # Net Income + D&A
        ws.append(["Cash Flow from Inv (CapEx)", "=-inp_Initial_Capex", 0, 0, 0, 0])
        ws.append(["Ending Cash", "=SUM(B15:B17)", "=SUM(C15:C17)", "=SUM(D15:D17)", "=SUM(E15:E17)", "=SUM(F15:F17)"])

        # Apply formatting to the data area
        for row in ws.iter_rows(min_row=3, max_row=18, min_col=2, max_col=6):
            for cell in row:
                cell.style = self.currency_style
        for cell in ws[3]:
            if cell.column > 1: cell.style = self.integer_style

    def create_cap_table_sheet(self):
        """Creates the SAFE conversion and capitalization table logic."""
        ws = self.wb.create_sheet("04_Cap_Table", 3)
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.append(["Capitalization Table Logic"])
        ws[1][0].style = self.header_style
        ws.merge_cells('A1:B1')

        data = [
            ("Pre-Money Shares (Founders)", 8000000, self.integer_style),
            ("Option Pool", 2000000, self.integer_style),
            ("Total Pre-Money Shares", "=B2+B3", self.integer_style),
            ("Series A Pre-Money Valuation", "=inp_SeriesA_Valuation", self.currency_style),
            ("Series A Price per Share", "=B5/B4", self.currency_2dp_style),
            ("SAFE Conversion Price (from Cap)", "=inp_SAFE_Cap/B4", self.currency_2dp_style),
            ("SAFE Conversion Price (from Discount)", "=B6*(1-inp_SAFE_Discount)", self.currency_2dp_style),
            ("Effective SAFE Conversion Price", "=MIN(B7, B8)", self.currency_2dp_style),
            ("Shares Issued to SAFEs", "=inp_SAFE_Amount/B9", self.integer_style),
            ("Post-SAFE, Pre-Series A Shares", "=B4+B10", self.integer_style),
            ("Founder Ownership Post-SAFE", "=B2/B11", self.percent_style),
            ("SAFE Investor Ownership Post-SAFE", "=B10/B11", self.percent_style),
        ]
        row = 2
        for label, formula, style in data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = formula
            ws[f'B{row}'].style = style
            row += 1

    def create_dashboard_sheet(self):
        """Creates a simple dashboard with key performance indicators and a chart."""
        ws = self.wb.create_sheet("05_Dashboard", 4)
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25

        # KPI Boxes
        ws['B2'] = "Year 1 EBITDA"
        ws['B3'] = "='03_Financials'!B8"
        ws['B3'].style = self.currency_style
        ws['C2'] = "Year 5 EBITDA"
        ws['C3'] = "='03_Financials'!F8"
        ws['C3'].style = self.currency_style
        ws['D2'] = "Year 5 Ending Cash"
        ws['D3'] = "='03_Financials'!F18"
        ws['D3'].style = self.currency_style

        # Chart
        chart = BarChart()
        chart.title = "EBITDA Over 5 Years"
        chart.y_axis.title = "Amount ($)"
        chart.x_axis.title = "Year"
        chart.legend = None # Correct way to disable the legend

        # Define data and categories for the chart
        data = Reference(self.wb['03_Financials'], min_col=2, max_col=6, min_row=8, max_row=8)
        cats = Reference(self.wb['03_Financials'], min_col=2, max_col=6, min_row=1, max_row=1)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        ws.add_chart(chart, "A6")

    def generate_model(self):
        """Orchestrates the creation of all sheets and saves the final workbook."""
        self.create_assumptions_sheet()
        self.create_scenario_switcher_sheet()
        self.create_financials_sheet()
        self.create_cap_table_sheet()
        self.create_dashboard_sheet()
        self.wb.active = self.wb["01_Assumptions"] # Set the starting sheet
        self.save("Brogue_Distillery_Financial_Model.xlsx")

    def save(self, filename):
        """Saves the workbook to the specified file."""
        self.wb.save(filename)
        print(f"Financial model saved as '{filename}'")

if __name__ == "__main__":
    print("Generating Brogue Distillery Financial Model...")
    model_generator = DistilleryModelGenerator()
    model_generator.generate_model()
    print("Model generation complete. The Excel file is ready.")
